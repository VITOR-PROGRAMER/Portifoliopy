import openpyxl
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
XLSX_SKILLS = BASE_DIR / "data" / "skills.xlsx"
SHEET_NAME = "Skills"
MAX_PROJETOS_POR_SKILL = 50


def skill_to_folder(skill: str) -> Path:
    cap = (skill or "").strip()
    return BASE_DIR / "Projetos" / (cap[:1].upper() + cap[1:].lower())


def contar_projetos(skill: str) -> int:
    pasta = skill_to_folder(skill)
    if not pasta.exists():
        return 0
    return sum(1 for p in pasta.iterdir() if p.is_dir())


def calcular_progressos(skills: list[str]) -> dict:
    out = {}
    for s in skills:
        c = contar_projetos(s)
        faltam = max(0, MAX_PROJETOS_POR_SKILL - c)
        percent = min(100, round((c / MAX_PROJETOS_POR_SKILL) * 100))
        out[s] = {"count": c, "faltam": faltam, "percent": percent}
    return out


def atualizar_planilha_por_projetos(
    skills: list[str], sobrescrever_value_com_percent=True
):
    if not XLSX_SKILLS.exists():
        print(f"[AVISO] Planilha não encontrada: {XLSX_SKILLS}")
        return

    wb = openpyxl.load_workbook(XLSX_SKILLS)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        print(f"[AVISO] Aba '{SHEET_NAME}' não encontrada em {XLSX_SKILLS}")
        return
    ws = wb[SHEET_NAME]

    headers_row = 1
    headers = {
        cell.value: idx
        for idx, cell in enumerate(ws[headers_row], start=1)
        if cell.value
    }

    def get_or_create_col(col_name: str) -> int:
        if col_name in headers:
            return headers[col_name]
        col_idx = ws.max_column + 1
        ws.cell(row=headers_row, column=col_idx, value=col_name)
        headers[col_name] = col_idx
        return col_idx

    col_skill = get_or_create_col("skill")
    col_value = get_or_create_col("value")
    col_proj = get_or_create_col("projects")
    col_missing = get_or_create_col("missing")
    col_percent = get_or_create_col("percent")

    skill_row = {}
    for r in range(2, ws.max_row + 1):
        key = ws.cell(row=r, column=col_skill).value
        if isinstance(key, str):
            skill_row[key.strip().lower()] = r

    prog = calcular_progressos(skills)

    for s in skills:
        s_key = s.strip().lower()
        row = skill_row.get(s_key)
        if row is None:
            row = ws.max_row + 1
            ws.cell(row=row, column=col_skill, value=s_key)
            skill_row[s_key] = row

        info = prog[s]
        ws.cell(row=row, column=col_proj, value=info["count"])
        ws.cell(row=row, column=col_missing, value=info["faltam"])
        ws.cell(row=row, column=col_percent, value=info["percent"])

        if sobrescrever_value_com_percent:
            ws.cell(row=row, column=col_value, value=info["percent"])

    wb.save(XLSX_SKILLS)
    wb.close()
    print("[OK] Planilha atualizada com projects / missing / percent.")
